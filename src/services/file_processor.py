"""File processing service for text extraction and image handling."""

import base64
import csv
import io
import mimetypes
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Optional

from src.utils.logger import logger


class FileCategory(str, Enum):
    """File category for processing strategy."""
    TEXT = "text"
    IMAGE = "image"
    DOCUMENT = "document"
    DATA = "data"
    CODE = "code"
    UNSUPPORTED = "unsupported"


@dataclass
class ProcessedFile:
    """Processed file result."""
    filename: str
    category: FileCategory
    mime_type: str
    size: int
    # For text/document/code/data: extracted text content
    text_content: Optional[str] = None
    # For images: base64 encoded data
    image_base64: Optional[str] = None
    # Error message if processing failed
    error: Optional[str] = None

    @property
    def is_success(self) -> bool:
        return self.error is None

    @property
    def has_text(self) -> bool:
        return self.text_content is not None and len(self.text_content) > 0

    @property
    def has_image(self) -> bool:
        return self.image_base64 is not None


# File type mappings
TEXT_EXTENSIONS = {".txt", ".md", ".rst", ".log", ".ini", ".cfg", ".conf"}
CODE_EXTENSIONS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".go", ".rs", ".rb", ".php", ".swift", ".kt", ".scala", ".sh",
    ".bash", ".zsh", ".ps1", ".sql", ".html", ".css", ".scss", ".sass",
    ".less", ".json", ".yaml", ".yml", ".xml", ".toml", ".env", ".gitignore",
    ".dockerfile", ".makefile", ".gradle", ".vue", ".svelte"
}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc"}
DATA_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}

# MIME type mappings for images (Vision API)
IMAGE_MIME_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".svg": "image/svg+xml",
}

# Max file sizes (bytes)
MAX_TEXT_SIZE = 10 * 1024 * 1024  # 10MB for text files
MAX_IMAGE_SIZE = 20 * 1024 * 1024  # 20MB for images
MAX_DOCUMENT_SIZE = 50 * 1024 * 1024  # 50MB for documents


def get_file_category(filename: str) -> FileCategory:
    """Determine file category from filename."""
    ext = Path(filename).suffix.lower()

    if ext in TEXT_EXTENSIONS:
        return FileCategory.TEXT
    elif ext in CODE_EXTENSIONS:
        return FileCategory.CODE
    elif ext in DOCUMENT_EXTENSIONS:
        return FileCategory.DOCUMENT
    elif ext in DATA_EXTENSIONS:
        return FileCategory.DATA
    elif ext in IMAGE_EXTENSIONS:
        return FileCategory.IMAGE
    else:
        return FileCategory.UNSUPPORTED


def get_mime_type(filename: str) -> str:
    """Get MIME type for a file."""
    ext = Path(filename).suffix.lower()
    if ext in IMAGE_MIME_TYPES:
        return IMAGE_MIME_TYPES[ext]
    mime_type, _ = mimetypes.guess_type(filename)
    return mime_type or "application/octet-stream"


class FileProcessor:
    """Process various file types for LLM context."""

    async def process(self, filename: str, content: bytes) -> ProcessedFile:
        """Process a file and extract content.

        Args:
            filename: Original filename
            content: File content as bytes

        Returns:
            ProcessedFile with extracted content
        """
        category = get_file_category(filename)
        mime_type = get_mime_type(filename)
        size = len(content)

        logger.info(f"Processing file: {filename} ({category.value}, {size} bytes)")

        try:
            if category == FileCategory.TEXT:
                return await self._process_text(filename, content, mime_type, size)
            elif category == FileCategory.CODE:
                return await self._process_code(filename, content, mime_type, size)
            elif category == FileCategory.DOCUMENT:
                return await self._process_document(filename, content, mime_type, size)
            elif category == FileCategory.DATA:
                return await self._process_data(filename, content, mime_type, size)
            elif category == FileCategory.IMAGE:
                return await self._process_image(filename, content, mime_type, size)
            else:
                return ProcessedFile(
                    filename=filename,
                    category=category,
                    mime_type=mime_type,
                    size=size,
                    error=f"Unsupported file type: {Path(filename).suffix}",
                )
        except Exception as e:
            logger.error(f"Failed to process {filename}: {e}")
            return ProcessedFile(
                filename=filename,
                category=category,
                mime_type=mime_type,
                size=size,
                error=str(e),
            )

    async def _process_text(
        self, filename: str, content: bytes, mime_type: str, size: int
    ) -> ProcessedFile:
        """Process plain text files."""
        if size > MAX_TEXT_SIZE:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.TEXT,
                mime_type=mime_type,
                size=size,
                error=f"File too large: {size} bytes (max: {MAX_TEXT_SIZE})",
            )

        text = self._decode_text(content)
        return ProcessedFile(
            filename=filename,
            category=FileCategory.TEXT,
            mime_type=mime_type,
            size=size,
            text_content=text,
        )

    async def _process_code(
        self, filename: str, content: bytes, mime_type: str, size: int
    ) -> ProcessedFile:
        """Process code files with syntax hints."""
        if size > MAX_TEXT_SIZE:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.CODE,
                mime_type=mime_type,
                size=size,
                error=f"File too large: {size} bytes (max: {MAX_TEXT_SIZE})",
            )

        text = self._decode_text(content)
        ext = Path(filename).suffix.lstrip(".")

        # Wrap code with language hint
        formatted = f"```{ext}\n{text}\n```"

        return ProcessedFile(
            filename=filename,
            category=FileCategory.CODE,
            mime_type=mime_type,
            size=size,
            text_content=formatted,
        )

    async def _process_document(
        self, filename: str, content: bytes, mime_type: str, size: int
    ) -> ProcessedFile:
        """Process PDF and Word documents."""
        if size > MAX_DOCUMENT_SIZE:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.DOCUMENT,
                mime_type=mime_type,
                size=size,
                error=f"File too large: {size} bytes (max: {MAX_DOCUMENT_SIZE})",
            )

        ext = Path(filename).suffix.lower()

        if ext == ".pdf":
            text = await self._extract_pdf(content)
        elif ext in {".docx", ".doc"}:
            text = await self._extract_docx(content)
        else:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.DOCUMENT,
                mime_type=mime_type,
                size=size,
                error=f"Unsupported document type: {ext}",
            )

        return ProcessedFile(
            filename=filename,
            category=FileCategory.DOCUMENT,
            mime_type=mime_type,
            size=size,
            text_content=text,
        )

    async def _process_data(
        self, filename: str, content: bytes, mime_type: str, size: int
    ) -> ProcessedFile:
        """Process CSV/Excel data files."""
        if size > MAX_TEXT_SIZE:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.DATA,
                mime_type=mime_type,
                size=size,
                error=f"File too large: {size} bytes (max: {MAX_TEXT_SIZE})",
            )

        ext = Path(filename).suffix.lower()

        if ext in {".csv", ".tsv"}:
            text = await self._extract_csv(content, ext)
        elif ext in {".xlsx", ".xls"}:
            text = await self._extract_excel(content)
        else:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.DATA,
                mime_type=mime_type,
                size=size,
                error=f"Unsupported data type: {ext}",
            )

        return ProcessedFile(
            filename=filename,
            category=FileCategory.DATA,
            mime_type=mime_type,
            size=size,
            text_content=text,
        )

    async def _process_image(
        self, filename: str, content: bytes, mime_type: str, size: int
    ) -> ProcessedFile:
        """Process image files for Vision API."""
        if size > MAX_IMAGE_SIZE:
            return ProcessedFile(
                filename=filename,
                category=FileCategory.IMAGE,
                mime_type=mime_type,
                size=size,
                error=f"Image too large: {size} bytes (max: {MAX_IMAGE_SIZE})",
            )

        # Encode to base64 for Vision API
        base64_data = base64.standard_b64encode(content).decode("utf-8")

        return ProcessedFile(
            filename=filename,
            category=FileCategory.IMAGE,
            mime_type=mime_type,
            size=size,
            image_base64=base64_data,
        )

    def _decode_text(self, content: bytes) -> str:
        """Decode bytes to text with encoding detection."""
        # Try common encodings
        encodings = ["utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"]

        for encoding in encodings:
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue

        # Last resort: decode with replacement
        return content.decode("utf-8", errors="replace")

    async def _extract_pdf(self, content: bytes) -> str:
        """Extract text from PDF."""
        try:
            import pdfplumber

            text_parts = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)

            return "\n\n".join(text_parts)
        except Exception as e:
            raise RuntimeError(f"PDF extraction failed: {e}")

    async def _extract_docx(self, content: bytes) -> str:
        """Extract text from DOCX."""
        try:
            from docx import Document

            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)
        except Exception as e:
            raise RuntimeError(f"DOCX extraction failed: {e}")

    async def _extract_csv(self, content: bytes, ext: str) -> str:
        """Extract text from CSV/TSV."""
        try:
            text = self._decode_text(content)
            delimiter = "\t" if ext == ".tsv" else ","

            # Parse CSV and format as markdown table
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)

            if not rows:
                return ""

            # Build markdown table
            lines = []
            header = rows[0]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")

            for row in rows[1:]:
                # Pad row if necessary
                padded = row + [""] * (len(header) - len(row))
                lines.append("| " + " | ".join(padded[:len(header)]) + " |")

            return "\n".join(lines)
        except Exception as e:
            raise RuntimeError(f"CSV extraction failed: {e}")

    async def _extract_excel(self, content: bytes) -> str:
        """Extract text from Excel."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            sheets_text = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    continue

                # Build markdown table for each sheet
                lines = [f"## Sheet: {sheet_name}"]

                # Find max columns
                max_cols = max(len(row) for row in rows) if rows else 0
                if max_cols == 0:
                    continue

                header = rows[0] if rows else []
                header = [str(c) if c is not None else "" for c in header]
                header = header + [""] * (max_cols - len(header))

                lines.append("| " + " | ".join(header) + " |")
                lines.append("| " + " | ".join(["---"] * max_cols) + " |")

                for row in rows[1:]:
                    cells = [str(c) if c is not None else "" for c in row]
                    cells = cells + [""] * (max_cols - len(cells))
                    lines.append("| " + " | ".join(cells) + " |")

                sheets_text.append("\n".join(lines))

            return "\n\n".join(sheets_text)
        except Exception as e:
            raise RuntimeError(f"Excel extraction failed: {e}")


# Global processor instance
_file_processor: Optional[FileProcessor] = None


def get_file_processor() -> FileProcessor:
    """Get or create file processor instance."""
    global _file_processor
    if _file_processor is None:
        _file_processor = FileProcessor()
    return _file_processor
